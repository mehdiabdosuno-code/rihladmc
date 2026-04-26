"""Excel Export — Formatted quotation spreadsheet.

Exports the full pricing grid with:
  - Cover sheet (project info)
  - Itinerary sheet (day by day)
  - Pricing sheet (multi-model, decomposition by category)
  - Comparison sheet (side by side models)
"""

import io
import os
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.database import get_db
from app.modules.itineraries.models import Itinerary, ItineraryDay
from app.modules.projects.models import Project
from app.modules.quotations.models import Quotation, QuotationLine
from app.shared.dependencies import require_auth

router = APIRouter(prefix="/export", tags=["excel-export"],
                   dependencies=[Depends(require_auth)])


def _col_letter(n: int) -> str:
    """Convert 0-based column index to Excel letter (0='A', 25='Z', 26='AA')."""
    result = ""
    while True:
        result = chr(n % 26 + 65) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


@router.get("/quotation/{project_id}", summary="Export quotation as Excel")
def export_quotation_excel(project_id: str, db: Session = Depends(get_db)):
    """Generate a formatted Excel workbook for the project quotation."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    # Load itinerary
    itin = db.execute(
        select(Itinerary).where(Itinerary.project_id == project_id)
    ).scalars().first()
    days = []
    if itin:
        days = db.execute(
            select(ItineraryDay)
            .where(ItineraryDay.itinerary_id == itin.id)
            .order_by(ItineraryDay.day_number)
        ).scalars().all()

    # Load quotation with lines
    quotation = db.execute(
        select(Quotation)
        .where(Quotation.project_id == project_id)
        .options(selectinload(Quotation.lines))
    ).scalars().first()

    wb = openpyxl.Workbook()

    # Styles
    navy = "1B2A4A"
    gold = "C5943A"
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    gold_fill = PatternFill(start_color=gold, end_color=gold, fill_type="solid")
    gold_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color=navy)
    subtitle_font = Font(bold=True, size=12, color=gold)
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0')
    )
    money_fmt = '#,##0.00 "€"'
    pct_fmt = '0.0%'

    # ── Sheet 1: Cover ───────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Proposition"
    ws_cover.sheet_properties.tabColor = navy

    ws_cover.column_dimensions['A'].width = 25
    ws_cover.column_dimensions['B'].width = 45

    ws_cover.merge_cells('A1:B1')
    ws_cover['A1'] = "S'TOURS — Destination Management Company"
    ws_cover['A1'].font = Font(bold=True, size=18, color=navy)

    ws_cover.merge_cells('A2:B2')
    ws_cover['A2'] = project.name
    ws_cover['A2'].font = subtitle_font

    info = [
        ("Client", project.client_name or "-"),
        ("Email", project.client_email or "-"),
        ("Destination", project.destination or "Maroc"),
        ("Durée", f"{project.duration_days or len(days)}J / {project.duration_nights or max(0, len(days)-1)}N"),
        ("Participants", f"{project.pax_count or '-'} personnes"),
        ("Type", str(project.project_type or "-")),
        ("Dates", project.travel_dates or "-"),
        ("Devise", project.currency or "EUR"),
        ("Statut", str(project.status)),
        ("Référence", project.reference or "-"),
        ("Date d'export", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
    ]

    for i, (label, value) in enumerate(info, start=4):
        ws_cover[f'A{i}'] = label
        ws_cover[f'A{i}'].font = Font(bold=True, color=navy)
        ws_cover[f'B{i}'] = value

    # ── Sheet 2: Itinerary ───────────────────────────────────────────
    ws_itin = wb.create_sheet("Programme")
    ws_itin.sheet_properties.tabColor = gold

    itin_headers = ["Jour", "Ville", "Titre", "Description", "Hôtel", "Distance (km)"]
    col_widths = [8, 18, 30, 50, 25, 15]

    for i, (h, w) in enumerate(zip(itin_headers, col_widths), start=1):
        cell = ws_itin.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws_itin.column_dimensions[get_column_letter(i)].width = w

    for row_idx, d in enumerate(days, start=2):
        ws_itin.cell(row=row_idx, column=1, value=d.day_number)
        ws_itin.cell(row=row_idx, column=2, value=d.city or "")
        ws_itin.cell(row=row_idx, column=3, value=d.title or "")
        ws_itin.cell(row=row_idx, column=4, value=(d.description or "")[:200])
        ws_itin.cell(row=row_idx, column=5, value=d.hotel or "")
        ws_itin.cell(row=row_idx, column=6, value=d.distance_km or 0)
        for c in range(1, 7):
            ws_itin.cell(row=row_idx, column=c).border = thin_border

    # ── Sheet 3: Pricing Detail ──────────────────────────────────────
    if quotation and quotation.lines:
        ws_price = wb.create_sheet("Cotation détaillée")
        ws_price.sheet_properties.tabColor = "2F855A"

        price_headers = ["Jour", "Catégorie", "Description", "Ville",
                         "Fournisseur", "Coût unitaire", "Quantité", "Unité", "Total"]
        price_widths = [8, 14, 35, 16, 20, 16, 10, 10, 16]

        for i, (h, w) in enumerate(zip(price_headers, price_widths), start=1):
            cell = ws_price.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws_price.column_dimensions[get_column_letter(i)].width = w

        # Group by category
        lines_sorted = sorted(quotation.lines,
                              key=lambda l: (l.day_number or 0, l.sort_order or 0))

        current_category = None
        row = 2
        for line in lines_sorted:
            cat = str(line.category) if line.category else "misc"
            if cat != current_category:
                # Category header row
                ws_price.merge_cells(f'A{row}:{get_column_letter(9)}{row}')
                ws_price.cell(row=row, column=1, value=cat.upper())
                ws_price.cell(row=row, column=1).font = gold_font
                ws_price.cell(row=row, column=1).fill = gold_fill
                current_category = cat
                row += 1

            ws_price.cell(row=row, column=1, value=line.day_number or "")
            ws_price.cell(row=row, column=2, value=cat)
            ws_price.cell(row=row, column=3, value=line.label or "")
            ws_price.cell(row=row, column=4, value=line.city or "")
            ws_price.cell(row=row, column=5, value=line.supplier or "")
            c_unit = ws_price.cell(row=row, column=6, value=float(line.unit_cost or 0))
            c_unit.number_format = money_fmt
            ws_price.cell(row=row, column=7, value=float(line.quantity or 1))
            ws_price.cell(row=row, column=8, value=line.unit or "")
            c_total = ws_price.cell(row=row, column=9, value=float(line.total_cost or 0))
            c_total.number_format = money_fmt
            for c in range(1, 10):
                ws_price.cell(row=row, column=c).border = thin_border
            row += 1

        # Totals
        row += 1
        ws_price.cell(row=row, column=8, value="TOTAL COÛT").font = Font(bold=True, color=navy)
        c_tc = ws_price.cell(row=row, column=9, value=float(quotation.total_cost or 0))
        c_tc.font = Font(bold=True, color=navy)
        c_tc.number_format = money_fmt

        row += 1
        ws_price.cell(row=row, column=8, value=f"Marge ({quotation.margin_pct}%)").font = Font(bold=True, color=gold)
        margin_val = float(quotation.total_selling or 0) - float(quotation.total_cost or 0)
        c_m = ws_price.cell(row=row, column=9, value=margin_val)
        c_m.font = Font(bold=True, color=gold)
        c_m.number_format = money_fmt

        row += 1
        ws_price.cell(row=row, column=8, value="TOTAL VENTE").font = Font(bold=True, size=13, color=navy)
        c_ts = ws_price.cell(row=row, column=9, value=float(quotation.total_selling or 0))
        c_ts.font = Font(bold=True, size=13, color=navy)
        c_ts.number_format = money_fmt

    # ── Sheet 4: Summary by category ─────────────────────────────────
    if quotation and quotation.lines:
        ws_summary = wb.create_sheet("Résumé catégories")
        ws_summary.sheet_properties.tabColor = "3182CE"

        sum_headers = ["Catégorie", "Nombre de lignes", "Total coût", "% du total"]
        sum_widths = [20, 18, 20, 15]
        for i, (h, w) in enumerate(zip(sum_headers, sum_widths), start=1):
            cell = ws_summary.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            ws_summary.column_dimensions[get_column_letter(i)].width = w

        cat_totals: dict[str, dict] = {}
        grand_total = 0
        for line in quotation.lines:
            cat = str(line.category) if line.category else "misc"
            if cat not in cat_totals:
                cat_totals[cat] = {"count": 0, "total": 0}
            cat_totals[cat]["count"] += 1
            cat_totals[cat]["total"] += float(line.total_cost or 0)
            grand_total += float(line.total_cost or 0)

        row = 2
        for cat, data in sorted(cat_totals.items(), key=lambda x: -x[1]["total"]):
            ws_summary.cell(row=row, column=1, value=cat.upper())
            ws_summary.cell(row=row, column=2, value=data["count"])
            c = ws_summary.cell(row=row, column=3, value=data["total"])
            c.number_format = money_fmt
            pct = data["total"] / grand_total if grand_total > 0 else 0
            c_p = ws_summary.cell(row=row, column=4, value=pct)
            c_p.number_format = pct_fmt
            row += 1

    # ── Save to buffer ───────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"STOURS_{project.reference or project_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/quotation-csv/{project_id}", summary="Export quotation as CSV")
def export_quotation_csv(project_id: str, db: Session = Depends(get_db)):
    """Lightweight CSV export of quotation lines."""
    import csv

    quotation = db.execute(
        select(Quotation)
        .where(Quotation.project_id == project_id)
        .options(selectinload(Quotation.lines))
    ).scalars().first()

    if not quotation:
        raise HTTPException(404, "No quotation found")

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(["Jour", "Catégorie", "Description", "Ville",
                     "Fournisseur", "Coût unitaire", "Quantité", "Unité", "Total"])

    for line in sorted(quotation.lines, key=lambda l: (l.day_number or 0, l.sort_order or 0)):
        writer.writerow([
            line.day_number or "",
            str(line.category) if line.category else "",
            line.label or "",
            line.city or "",
            line.supplier or "",
            float(line.unit_cost or 0),
            float(line.quantity or 1),
            line.unit or "",
            float(line.total_cost or 0),
        ])

    buffer.seek(0)
    return StreamingResponse(
        io.BytesIO(buffer.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="cotation_{project_id[:8]}.csv"'},
    )
